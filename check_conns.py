import math
import os, json
import openpyxl
import numpy as np
import numpy.lib.recfunctions as recfunctions
from matplotlib import pyplot

if __name__ == '__main__':
    conns = {}
    root = 'D:/Celegans-locomotion/data/new_example_all/'
    # for f in os.listdir(root):
    #     print(f)
    #     if os.path.isdir(root + f):
    #         conn_name, value = f.split("_")
    #         if conn_name not in conns.keys():
    #             conns[conn_name] = {"food_left": [], "food_none": [], "food_right": []}
    #         for sub_folder in ["food_left", "food_none", "food_right"]:
    #             data_raw = np.genfromtxt(root + f + "/" + sub_folder + "/body.dat", delimiter=' ', dtype=None)
    #             data_raw = data_raw[:, 1:]
    #             n_tstep = data_raw.shape[0]
    #             n_seg = data_raw.shape[1] // 3
    #             data_raw = np.reshape(data_raw, (
    #             n_tstep, n_seg, len(np.dtype([('x', 'f8'), ('y', 'f8'), ('phi', 'f8')]))))  # type: ignore
    #             data_raw = recfunctions.unstructured_to_structured(
    #                 data_raw,
    #                 dtype=np.dtype([('x', 'f8'), ('y', 'f8'), ('phi', 'f8')]),
    #             )
    #             conns[conn_name][sub_folder].append(
    #                 [value, data_raw['x'][-1, 0], data_raw['y'][-1, 0], data_raw['phi'][-1, 0]])

    # with open("dataset", 'w') as fout:
    #     json.dump(conns, fout, indent=4)
    with open("dataset", 'r', encoding='utf-8') as fin:
        conns: dict = json.load(fin)

    # for key, value in conns.items():
    #     wb = openpyxl.Workbook()
    #     for sub_folder, params in value.items():
    #         ws = wb.create_sheet(sub_folder)
    #         for row in range(len(params)):
    #             for col in range(len(params[0])):
    #                 # print(row, col, params[row][col])
    #                 ws.cell(row=row+1, column=col+1).value = params[row][col]
    #     wb.save(key+'.xlsx')
    for key, value in conns.items():
        for sub_folder, params in value.items():
            if sub_folder == "food_left":
                food_pos = [-0.03, 0.05]
            elif sub_folder == "food_none":
                food_pos = [0.0, 0.05]
            elif sub_folder == "food_right":
                food_pos = [0.03, 0.05]
            params = np.array(params,dtype='float64')

            distance = (params[:, 1]-food_pos[0])*(params[:, 1]-food_pos[0])+(params[:, 2]-food_pos[0])*(params[:, 2]-food_pos[0])
            pyplot.plot(params[:, 0], distance**0.5, "*-")
            title = key+"-"+sub_folder
            pyplot.title(title)
            pyplot.savefig("D:/Celegans-locomotion/data/new_sweep_conns/"+title+".jpeg")
            print(title)
            pyplot.cla()
            # pyplot.show()

